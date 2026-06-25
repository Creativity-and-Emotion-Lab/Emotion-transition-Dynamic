import math
from pathlib import Path

import numpy as np
import pandas as pd
import pingouin as pg
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


input_path = Path(
    r"E:\Analysis\EMA\brain_match_newdata.xlsx"
)

sheet_name = 0

output_path = input_path.parent / "DMN_FPCN_null_results.xlsx"


brain_vars = [
    "Net6",
    "Net7",
]

emotion_vars = [
    "PAPR_L",
    "PAPR_M",
    "PAPR_H",
    "NAPR_L",
    "NAPR_M",
    "NAPR_H",
    "PA_Dens",
    "NA_Dens",
    "Joint_Dens",
]


alpha = 0.05

target_power = 0.80

alternative = "two-sided"


def interpret_bf01(bf01: float) -> str:

    if not np.isfinite(bf01):
        return "Not available"

    if bf01 >= 10:
        return "Strong evidence for H0"

    if bf01 >= 3:
        return "Moderate evidence for H0"

    if bf01 > 1:
        return "Weak evidence for H0"

    if math.isclose(bf01, 1.0):
        return "Equal evidence for H0 and H1"

    if bf01 > 1 / 3:
        return "Inconclusive evidence"

    if bf01 > 0.1:
        return "Moderate evidence for H1"

    return "Strong evidence for H1"


def convert_bf10(value) -> float:

    if isinstance(value, str):
        value = value.strip()

        if value.lower() in {"inf", "+inf", "infinity"}:
            return math.inf

    return float(value)


def extract_ci(ci_value) -> tuple[float, float]:

    if isinstance(ci_value, (list, tuple, np.ndarray)):
        ci_array = np.asarray(ci_value, dtype=float).flatten()

        if len(ci_array) != 2:
            raise ValueError(
                f"Could not parse CI value: {ci_value}"
            )

        return float(ci_array[0]), float(ci_array[1])

    if isinstance(ci_value, str):
        cleaned = (
            ci_value
            .replace("[", "")
            .replace("]", "")
            .replace("(", "")
            .replace(")", "")
            .replace(",", " ")
        )

        values = [
            float(item)
            for item in cleaned.split()
            if item.strip()
        ]

        if len(values) != 2:
            raise ValueError(
                f"Could not parse CI value: {ci_value}"
            )

        return values[0], values[1]

    raise TypeError(
        f"Unsupported CI value type: {type(ci_value).__name__}"
    )


if not input_path.exists():
    raise FileNotFoundError(
        f"Input file not found:\n{input_path}"
    )


df = pd.read_excel(
    input_path,
    sheet_name=sheet_name,
)

excel_row_to_delete = 28
dataframe_index = excel_row_to_delete - 2

if dataframe_index not in df.index:
    raise IndexError(
        f"Excel row {excel_row_to_delete} does not exist. "
        f"{len(df)} data rows found."
    )

df = df.drop(index=dataframe_index).reset_index(drop=True)

print(f"Deleted Excel row {excel_row_to_delete}")

df.columns = [
    str(col).strip()
    for col in df.columns
]

required_cols = brain_vars + emotion_vars

missing_cols = [
    col
    for col in required_cols
    if col not in df.columns
]

if missing_cols:
    raise ValueError(
        "The following variables were not found in the Excel file:\n\n"
        + "\n".join(missing_cols)
        + "\n\nAvailable columns:\n\n"
        + "\n".join(map(str, df.columns))
    )


for col in required_cols:

    try:
        df[col] = pd.to_numeric(
            df[col],
            errors="raise",
        )

    except (ValueError, TypeError) as error:

        converted = pd.to_numeric(
            df[col],
            errors="coerce",
        )

        invalid_mask = (
            df[col].notna()
            & converted.isna()
        )

        invalid_rows = df.index[invalid_mask].tolist()
        invalid_values = (
            df.loc[invalid_mask, col]
            .astype(str)
            .tolist()
        )

        details = []

        for row_index, value in zip(
            invalid_rows,
            invalid_values,
        ):
            excel_row = row_index + 2

            details.append(
                f"Excel row {excel_row}: {value!r}"
            )

        raise ValueError(
            f"Variable {col!r} contains non-numeric values:\n"
            + "\n".join(details)
        ) from error


df[required_cols] = df[required_cols].astype(float)


missing_counts = df[required_cols].isna().sum()

columns_with_missing = missing_counts[
    missing_counts > 0
]

if not columns_with_missing.empty:
    missing_message = "\n".join(
        f"{col}: {count} missing values"
        for col, count in columns_with_missing.items()
    )

    raise ValueError(
        "Missing values were detected in the analysis variables:\n\n"
        + missing_message
    )


for col in required_cols:

    nonfinite_mask = ~np.isfinite(
        df[col].to_numpy(dtype=float)
    )

    if nonfinite_mask.any():
        row_numbers = (
            df.index[nonfinite_mask] + 2
        ).tolist()

        raise ValueError(
            f"Variable {col!r} contains infinite or invalid values. "
            f"Excel rows: {row_numbers}"
        )


constant_cols = [
    col
    for col in required_cols
    if df[col].nunique() < 2
]

if constant_cols:
    raise ValueError(
        "The following variables are constant and cannot be correlated:\n\n"
        + "\n".join(constant_cols)
    )


results = []

total_tests = len(brain_vars) * len(emotion_vars)

print("=" * 60)
print(f"Running {total_tests} correlations")
print(f"Pingouin version: {pg.__version__}")
print("=" * 60)


for brain_var in brain_vars:

    for emotion_var in emotion_vars:

        x = df[brain_var].to_numpy(dtype=float)
        y = df[emotion_var].to_numpy(dtype=float)

        corr_result = pg.corr(
            x=x,
            y=y,
            method="pearson",
            alternative=alternative,
        )

        row = corr_result.iloc[0]

        n = int(row["n"])
        r = float(row["r"])
        p_value = float(row["p_val"])

        ci_columns = [
            col
            for col in corr_result.columns
            if str(col).upper().startswith("CI95")
        ]

        if not ci_columns:
            raise KeyError(
                "No 95% confidence interval column was found in the Pingouin output.\n"
                f"Current columns: {list(corr_result.columns)}"
            )

        ci_value = row[ci_columns[0]]

        ci_low, ci_high = extract_ci(ci_value)

        bf10 = convert_bf10(row["BF10"])

        if math.isinf(bf10):
            bf01 = 0.0

        elif bf10 > 0:
            bf01 = 1.0 / bf10

        else:
            bf01 = math.nan

        if "power" in corr_result.columns:
            achieved_power = float(row["power"])
        else:
            achieved_power = pg.power_corr(
                r=abs(r),
                n=n,
                alpha=alpha,
                power=None,
                alternative=alternative,
            )

        detectable_r = pg.power_corr(
            r=None,
            n=n,
            power=target_power,
            alpha=alpha,
            alternative=alternative,
        )

        results.append(
            {
                "Brain variable": brain_var,
                "Emotion variable": emotion_var,
                "n": n,
                "r": r,
                "95% CI lower": ci_low,
                "95% CI upper": ci_high,
                "p": p_value,
                "BF10": bf10,
                "BF01": bf01,
                "BF01 interpretation": interpret_bf01(
                    bf01
                ),
                "Achieved power": float(
                    achieved_power
                ),
                "Minimum detectable |r|": abs(
                    float(detectable_r)
                ),
                "Target power": target_power,
                "Alpha": alpha,
            }
        )

        print(
            f"{brain_var} x {emotion_var}: "
            f"n={n}, "
            f"r={r:.3f}, "
            f"p={p_value:.4f}, "
            f"BF01={bf01:.3f}"
        )


results_df = pd.DataFrame(results)


if results_df.empty:
    raise RuntimeError(
        "No correlation results were generated."
    )


reject_fdr, p_fdr = pg.multicomp(
    results_df["p"].to_numpy(dtype=float),
    alpha=alpha,
    method="fdr_bh",
)

results_df["p_FDR"] = p_fdr

results_df["Significant after FDR"] = reject_fdr

results_df["Raw p significant"] = (
    results_df["p"] < alpha
)


result_columns = [
    "Brain variable",
    "Emotion variable",
    "n",
    "r",
    "95% CI lower",
    "95% CI upper",
    "p",
    "Raw p significant",
    "p_FDR",
    "Significant after FDR",
    "BF10",
    "BF01",
    "BF01 interpretation",
    "Achieved power",
    "Minimum detectable |r|",
    "Target power",
    "Alpha",
]

results_df = results_df[result_columns]


null_results_df = results_df.loc[
    results_df["p"] >= alpha
].copy()


null_results_df = null_results_df.sort_values(
    by="BF01",
    ascending=False,
)


unique_ns = sorted(
    results_df["n"].unique()
)

power_rows = []

for n in unique_ns:

    detectable_r = pg.power_corr(
        r=None,
        n=int(n),
        power=target_power,
        alpha=alpha,
        alternative=alternative,
    )

    power_rows.append(
        {
            "n": int(n),
            "Alpha": alpha,
            "Target power": target_power,
            "Minimum detectable |r|": abs(
                float(detectable_r)
            ),
            "Interpretation": (
                f"With n={int(n)}, the study had "
                f"{target_power:.0%} power to detect "
                f"|r| >= {abs(float(detectable_r)):.3f} "
                f"at alpha={alpha:.2f}."
            ),
        }
    )

power_df = pd.DataFrame(power_rows)


variable_check_rows = []

for col in required_cols:

    variable_check_rows.append(
        {
            "Variable": col,
            "Variable type": (
                "Brain variable"
                if col in brain_vars
                else "Emotion variable"
            ),
            "Data type": str(df[col].dtype),
            "n": int(df[col].count()),
            "Missing values": int(
                df[col].isna().sum()
            ),
            "Unique values": int(
                df[col].nunique()
            ),
            "Mean": float(df[col].mean()),
            "SD": float(df[col].std(ddof=1)),
            "Minimum": float(df[col].min()),
            "Maximum": float(df[col].max()),
        }
    )

variable_check_df = pd.DataFrame(
    variable_check_rows
)


with pd.ExcelWriter(
    output_path,
    engine="openpyxl",
) as writer:

    results_df.to_excel(
        writer,
        sheet_name="All_correlations",
        index=False,
    )

    null_results_df.to_excel(
        writer,
        sheet_name="Nonsignificant_results",
        index=False,
    )

    power_df.to_excel(
        writer,
        sheet_name="Sensitivity_power",
        index=False,
    )

    variable_check_df.to_excel(
        writer,
        sheet_name="Variable_check",
        index=False,
    )


workbook = load_workbook(output_path)


for worksheet in workbook.worksheets:

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in worksheet.columns:

        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells:

            if cell.value is not None:
                cell_length = len(str(cell.value))
                max_length = max(
                    max_length,
                    cell_length,
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 45)


workbook.save(output_path)


print("=" * 60)
print("Analysis completed")
print(f"Total correlations: {len(results_df)}")
print(
    f"Nonsignificant raw p-value results: "
    f"{len(null_results_df)}"
)
print(f"Output file:\n{output_path}")
print("=" * 60)
